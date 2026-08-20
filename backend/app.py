# -*- coding: utf-8 -*-
import sys
import os
import json
from pathlib import Path
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import openpyxl

app = Flask(__name__)
CORS(app)

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = Path(__file__).parent / "config.json"

DOCUMENT_UPDATE_DIR = ROOT / "scripts" / "document_update"

# ── 載入配置 ──────────────────────────────────────────────────────────────────
def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

config = load_config()

def get_config():
    """每次呼叫時重新讀取 config，讓修改 config.json 後不需重啟 Flask"""
    return load_config()

# ── sys.path：讓 Flask 程序可以直接 import script 模組 ────────────────────────
if str(DOCUMENT_UPDATE_DIR) not in sys.path:
    sys.path.insert(0, str(DOCUMENT_UPDATE_DIR))

from core.report_generator import main as _run_document_update
from core.snapshot_utils import take_snapshot, SNAPSHOT_FILENAME

# ── 回應 helper ───────────────────────────────────────────────────────────────
def ok(**kwargs):
    return jsonify({"ok": True, **kwargs})

def err(message: str, status: int = 400, **kwargs):
    return jsonify({"ok": False, "message": message, **kwargs}), status

# ── 路由 ──────────────────────────────────────────────────────────────────────
@app.get("/")
def serve_ui():
    html_path = ROOT / "web" / "IR小工具.html"
    return send_file(str(html_path))


@app.get("/health")
def health():
    cfg = get_config()
    return ok(scripts=list(cfg["scripts"].keys()))


@app.get("/scripts")
def list_scripts():
    return ok(scripts=config["scripts"])


@app.post("/run/<name>")
def run_script(name: str):
    if name not in config["scripts"]:
        return err(f"腳本 '{name}' 不在允許清單中", 404)

    if name == "script1":
        excel_path = str(DOCUMENT_UPDATE_DIR / "input" / "data input.xlsx")
        if not os.path.exists(excel_path):
            return err(f"Excel 檔案不存在: {excel_path}", 404)

        VALID_REPORTS = {"transcript_en", "transcript_zh", "management_report"}
        data = request.get_json(silent=True) or {}
        selected_reports = data.get("selected_reports") or list(VALID_REPORTS)
        invalid = set(selected_reports) - VALID_REPORTS
        if invalid:
            return err(f"不支援的報告類型: {sorted(invalid)}")

        prev_dir = os.getcwd()
        try:
            os.chdir(str(DOCUMENT_UPDATE_DIR))
            _run_document_update(excel_path, selected_reports)
            return ok(output="文件更新處理完成！")
        except Exception as e:
            return err(f"執行錯誤: {e}", 500)
        finally:
            os.chdir(prev_dir)

    return err(f"腳本 '{name}' 尚未支援直接呼叫", 501)


@app.get("/excel-path")
def get_excel_path():
    excel_path = DOCUMENT_UPDATE_DIR / "input" / "data input.xlsx"
    return ok(path=str(excel_path), exists=excel_path.exists())


@app.post("/upload-excel")
def upload_excel():
    """接收上傳的 Excel 檔案並存到 input/data input.xlsx（給 n8n 之類的外部呼叫者用，
    取代直接在伺服器上用檔案總管放檔案）"""
    if "file" not in request.files:
        return err("沒有收到檔案（form-data 欄位需命名為 'file'）")

    upload = request.files["file"]
    if not upload.filename:
        return err("檔案名稱是空的")
    if not upload.filename.lower().endswith((".xlsx", ".xls")):
        return err("檔案格式錯誤，只接受 .xlsx / .xls")

    excel_path = DOCUMENT_UPDATE_DIR / "input" / "data input.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    upload.save(str(excel_path))
    return ok(message="Excel 已上傳", path=str(excel_path))


@app.post("/open-excel")
def open_excel():
    excel_path = DOCUMENT_UPDATE_DIR / "input" / "data input.xlsx"
    if not excel_path.exists():
        return err(f"Excel 檔案不存在: {excel_path}", 404)
    try:
        os.startfile(str(excel_path))
        return ok(message="已開啟 Excel")
    except Exception as e:
        return err(f"開啟失敗: {e}", 500)


@app.post("/update-config")
def update_document_config():
    data = request.get_json() or {}
    quarter = data.get("quarter")
    year = data.get("year")
    event_date = data.get("event_date")

    if not all([quarter, year, event_date]):
        return err("缺少必要參數 (quarter, year, event_date)")

    config_path = DOCUMENT_UPDATE_DIR / "configs" / "quarter_config.json"
    if not config_path.exists():
        return err("配置文件不存在", 404)

    with open(config_path, "r", encoding="utf-8") as f:
        doc_config = json.load(f)

    doc_config.update({"this_quarter": quarter, "this_year": year, "event_date": event_date})

    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(doc_config, f, indent=2, ensure_ascii=False)

    # ── 取得當前 Excel 快照（用於下一季的 stale-data 偵測）─────────────────────
    snapshot_taken = False
    snapshot_message = ""
    try:
        data_config_path = DOCUMENT_UPDATE_DIR / "configs" / "data_config.json"
        excel_path = DOCUMENT_UPDATE_DIR / "input" / "data input.xlsx"
        output_dir = DOCUMENT_UPDATE_DIR / "output"
        snapshot_path = str(output_dir / SNAPSHOT_FILENAME)

        if excel_path.exists() and data_config_path.exists():
            with open(data_config_path, "r", encoding="utf-8") as f:
                data_cfg = json.load(f)
            sheet_names = list(data_cfg["sheet_mapping"].keys())
            output_dir.mkdir(exist_ok=True)
            take_snapshot(str(excel_path), sheet_names, quarter, str(year), snapshot_path)
            snapshot_taken = True
            snapshot_message = f"已對 {year} {quarter} 取得快照（{len(sheet_names)} 個 sheet）"
        else:
            snapshot_message = "Excel 或 data_config 不存在，跳過快照"
    except Exception as _e:
        snapshot_message = f"快照失敗（不影響主流程）: {_e}"

    return ok(message="配置更新成功", snapshot=snapshot_taken, snapshot_detail=snapshot_message)


# ── 講稿文字 sheet 的讀取與寫入 ───────────────────────────────────────────────
SPEECH_SHEETS = {"opening_remarks", "future_outlook", "chairman_remarks"}


@app.get("/get-speech/<sheet_name>")
def get_speech_content(sheet_name: str):
    if sheet_name not in SPEECH_SHEETS:
        return err(f"不支援的 sheet: {sheet_name}（允許: {sorted(SPEECH_SHEETS)}）", 400)

    excel_path = DOCUMENT_UPDATE_DIR / "input" / "data input.xlsx"
    if not excel_path.exists():
        return err("找不到 data input.xlsx", 404)

    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return err(f"Excel 中找不到 sheet: {sheet_name}", 404)
        ws = wb[sheet_name]

        # Read header row to find column indices
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        def col_idx(name):
            try: return header.index(name)
            except ValueError: return None

        idx_en    = col_idx("content_en")
        idx_zh    = col_idx("content_zh")
        idx_title = col_idx("title")

        def cell_str(row_cells, idx):
            if idx is None or idx >= len(row_cells):
                return ""
            v = row_cells[idx].value
            return "" if v is None else str(v)

        rows = []
        for row_cells in ws.iter_rows(min_row=2):
            key_val = row_cells[0].value
            if key_val is None:
                continue
            entry = {"key": str(key_val)}
            entry["content_en"]  = cell_str(row_cells, idx_en)
            entry["content_zh"]  = cell_str(row_cells, idx_zh)
            entry["title"]       = cell_str(row_cells, idx_title)
            rows.append(entry)

        wb.close()
        return ok(sheet=sheet_name, rows=rows)
    except Exception as e:
        return err(f"讀取失敗: {e}", 500)


@app.post("/save-speech")
def save_speech_content():
    data = request.get_json() or {}
    sheet_name = data.get("sheet_name")
    rows = data.get("rows", [])

    if sheet_name not in SPEECH_SHEETS:
        return err(f"不支援的 sheet: {sheet_name}", 400)
    if not rows:
        return err("沒有資料", 400)

    excel_path = DOCUMENT_UPDATE_DIR / "input" / "data input.xlsx"
    if not excel_path.exists():
        return err("找不到 data input.xlsx", 404)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path))
        if sheet_name not in wb.sheetnames:
            return err(f"Sheet '{sheet_name}' 不存在於 Excel 中", 404)

        ws = wb[sheet_name]

        # header row → column index map
        header = {}
        for cell in ws[1]:
            if cell.value is not None:
                header[str(cell.value)] = cell.column

        # first column (index_col=0) → row number map
        key_to_row = {}
        for r in ws.iter_rows(min_row=2):
            first = r[0].value
            if first is not None:
                key_to_row[str(first)] = r[0].row

        updated = 0
        for entry in rows:
            key = str(entry.get("key", ""))
            if key not in key_to_row:
                continue
            row_num = key_to_row[key]
            for col_name in ["content_en", "content_zh", "title"]:
                if col_name in entry and col_name in header:
                    ws.cell(row=row_num, column=header[col_name]).value = entry[col_name]
            updated += 1

        wb.save(str(excel_path))
        return ok(message=f"已更新 {updated} 筆", sheet=sheet_name, updated=updated)
    except Exception as e:
        return err(f"儲存失敗: {e}", 500)

if __name__ == "__main__":
    srv = config["server"]
    print(f"Starting Flask server on {srv['host']}:{srv['port']} ...")
    app.run(host=srv["host"], port=srv["port"], debug=srv["debug"])
